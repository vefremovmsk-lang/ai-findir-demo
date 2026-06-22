# -*- coding: utf-8 -*-
"""
make_sample_data — генератор СИНТЕТИЧЕСКИХ демо-данных для публичного репо AI-Findir.

Создаёт вымышленную компанию, отделы, ФИО и цифры в ТОМ ЖЕ формате, что читает движок
(00-слой + главный лист с 5 секциями + Tier-2 Договора/КП для экологии). Никаких реальных
данных. Запуск: python make_sample_data.py  → zero_layer.xlsx + raw/*.xlsx рядом.
"""
import os, random
from datetime import date
import openpyxl

random.seed(42)
HERE = os.path.dirname(os.path.abspath(__file__))
CTRL = date(2026, 6, 19)

FAKE_MGRS = ["Соколов", "Зимин", "Лебедев", "Орлова"]
FAKE_EXEC = ["Ковалёв", "Тихонова", "Рябов", "Зорин"]
FAKE_CLIENTS = ["ООО «Вектор-Пром»", "АО «Северсталь-Тест»", "ООО «ТехноЛайн»",
                "АО «Гранит»", "ООО «Метрополь»", "АО «Сибэнерго»", "ООО «Альфа-Снаб»"]

def m(v): return int(round(v))

def main_sheet(ws, title, managers, execs, scale):
    ws["A1"] = f"{title} — рабочий лист (ДЕМО · синтетические данные)"
    ws["A3"] = "Дата заполнения"; ws["B3"] = CTRL.isoformat()
    months = ["Апрель", "Май", "Июнь"]
    # --- 1. свод ---
    ws["A10"] = "1. Свод Q2 по месяцам"
    hdr = ["Месяц","План контрактации","Факт контрактации","Прогноз контрактации",
           "% выполнения","План актирования","Факт выручки по актам","Прогноз актирования",
           "Портфель в работе","Расходы направления","Опер. финрез","Рентабельность",
           "Эффект инициатив (продажи)","Эффект инициатив (исполн.)"]
    for c,h in enumerate(hdr,1): ws.cell(11,c,h)
    q = {k:0 for k in ("cp","cf","ap","af","exp","port")}
    r = 12
    for i,mo in enumerate(months):
        cp=m(scale*random.uniform(1.5,2.2)*1e6); cf=m(cp*random.uniform(0.5,0.9))
        ap=m(scale*random.uniform(1.4,2.0)*1e6); af=m(ap*random.uniform(0.45,0.85))
        exp=m(scale*random.uniform(1.5,1.9)*1e6); port=m(scale*random.uniform(0.5,2.0)*1e6)
        fin=af-exp
        ws.cell(r,1,mo); ws.cell(r,2,cp); ws.cell(r,3,cf); ws.cell(r,4,0)
        ws.cell(r,5,round(cf/cp,4)); ws.cell(r,6,ap); ws.cell(r,7,af); ws.cell(r,8,0)
        ws.cell(r,9,port); ws.cell(r,10,exp); ws.cell(r,11,fin)
        ws.cell(r,12,round(fin/af,4) if af else 0)
        for k,v in (("cp",cp),("cf",cf),("ap",ap),("af",af),("exp",exp),("port",port)): q[k]+=v
        r+=1
    fin=q["af"]-q["exp"]
    ws.cell(r,1,"Q2 Итого"); ws.cell(r,2,q["cp"]); ws.cell(r,3,q["cf"]); ws.cell(r,4,0)
    ws.cell(r,5,round(q["cf"]/q["cp"],4)); ws.cell(r,6,q["ap"]); ws.cell(r,7,q["af"]); ws.cell(r,8,0)
    ws.cell(r,9,q["port"]); ws.cell(r,10,q["exp"]); ws.cell(r,11,fin)
    ws.cell(r,12,round(fin/q["af"],4) if q["af"] else 0)
    # --- 2. менеджеры ---
    r+=2; ws.cell(r,1,"2. Продажи и контрактация по менеджерам"); r+=1
    for c,h in enumerate(["Месяц","Менеджер","План","Факт","Прогноз","На подписании","Отклонение","%","Сделок","Эффект","Статус"],1): ws.cell(r,c,h)
    r+=1
    for mo in months:
        for nm in managers:
            pl=m(scale*random.uniform(0.3,0.8)*1e6); fa=m(pl*random.uniform(0.2,1.3))
            ws.cell(r,1,mo); ws.cell(r,2,nm); ws.cell(r,3,pl); ws.cell(r,4,fa)
            ws.cell(r,5,0); ws.cell(r,6,0); ws.cell(r,7,fa-pl); ws.cell(r,8,round(fa/pl,3))
            ws.cell(r,9,random.randint(1,6)); ws.cell(r,10,0); ws.cell(r,11,"в работе"); r+=1
    # --- 3. исполнители ---
    r+=1; ws.cell(r,1,"3. Исполнение и актирование по исполнителям"); r+=1
    for c,h in enumerate(["Месяц","Исполнитель","План","Факт","Прогноз","Портфель","Отклонение","%","Актов","Эффект","Статус"],1): ws.cell(r,c,h)
    r+=1
    for mo in months:
        for nm in execs:
            pl=m(scale*random.uniform(0.4,0.9)*1e6); fa=m(pl*random.uniform(0.3,1.1))
            ws.cell(r,1,mo); ws.cell(r,2,nm); ws.cell(r,3,pl); ws.cell(r,4,fa)
            ws.cell(r,5,0); ws.cell(r,6,m(pl*0.5)); ws.cell(r,7,fa-pl); ws.cell(r,8,round(fa/pl,3))
            ws.cell(r,9,random.randint(3,12)); ws.cell(r,10,0); ws.cell(r,11,"в работе"); r+=1
    # --- 4. расходы ---
    r+=1; ws.cell(r,1,"4. Расходы и операционный финрезультат"); r+=1
    for c,h in enumerate(["Месяц","Распр. админ","Прямой ФОТ","Коммерч. ФОТ","Прочие","Общехоз","Расходы всего","Выручка","Финрез","Рентаб."],1): ws.cell(r,c,h)
    r+=1
    for mo in months:
        tot=m(scale*random.uniform(1.5,1.9)*1e6)
        ws.cell(r,1,mo); ws.cell(r,2,m(tot*0.2)); ws.cell(r,3,m(tot*0.35)); ws.cell(r,4,m(tot*0.3))
        ws.cell(r,5,m(tot*0.1)); ws.cell(r,6,m(tot*0.05)); ws.cell(r,7,tot); r+=1
    # --- 5. недельные действия ---
    r+=1; ws.cell(r,1,"5. Недельные действия и инициативы"); r+=1
    for c,h in enumerate(["Месяц","Менеджер","Инициатива","Ожид. контракт","Факт","Исполнитель","Инициатива","Ожид. акт","Факт"],1): ws.cell(r,c,h)
    r+=1
    inits=["Партнёрская программа с дистрибьютором","Заключение договора с новым клиентом",
           "Допродажа сопровождения базе","Тендерная заявка по 44-ФЗ"]
    for i,nm in enumerate(managers):
        ws.cell(r,1,"Июнь"); ws.cell(r,2,nm); ws.cell(r,3,random.choice(inits))
        ws.cell(r,4,m(scale*random.uniform(0.3,1.0)*1e6)); ws.cell(r,5,0); r+=1

def deals_sheet(ws, managers):
    ws["A1"]="Июнь"; ws["A2"]="Компания:"; ws["C2"]="Сумма с НДС"; ws["D2"]="Сумма без НДС"; ws["E2"]="Менеджер"
    r=3
    for cl in FAKE_CLIENTS:
        s=random.randint(80,600)*1000
        ws.cell(r,1,cl); ws.cell(r,3,s); ws.cell(r,4,m(s/1.2)); ws.cell(r,5,random.choice(managers)); r+=1

def kp_sheet(ws, managers):
    ws["B1"]="Демо"; ws["D1"]="Продажи"
    # строка label-1 с менеджерами (план в колонке, факт в col+1), начиная с col 4
    c=4
    for nm in managers:
        ws.cell(2,c,nm); c+=2
    ws["A3"]="Количество КП"
    for c2 in range(4,4+2*len(managers),2):
        ws.cell(3,c2,"План по КП"); ws.cell(3,c2+1,"Факт по КП")
    r=4
    for mo in ["Апрель","Май","Июнь"]:
        ws.cell(r,1,mo); c=4
        for nm in managers:
            ws.cell(r,c,random.randint(15,40)); ws.cell(r,c+1,random.randint(5,35)); c+=2
        r+=1

DEPTS = [
    ("ecology_q2_2026_0619.xlsx", "Экология", "Экологические услуги", True, 1.0),
    ("product_certification_q2_2026_0619.xlsx", "Сертификация продукции", "Сертификация продукции", False, 1.6),
    ("training_center_q2_2026_0619.xlsx", "Учебный центр", "Учебный центр", False, 0.6),
    ("qms_q2_2026_0619.xlsx", "ОС СМ", "Сертификация СМК", False, 1.1),
]
for fn, sheet1, title, has_t2, scale in DEPTS:
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=sheet1
    mgrs=random.sample(FAKE_MGRS,3); exs=random.sample(FAKE_EXEC,2)
    main_sheet(ws,title,mgrs,exs,scale)
    if has_t2:
        deals_sheet(wb.create_sheet("Договора"),mgrs)
        kp_sheet(wb.create_sheet("КП и Звонки"),mgrs)
    wb.save(os.path.join(HERE,"raw",fn))
    print("xlsx ->",fn)

# ---- 00-слой ----
wb=openpyxl.Workbook(); info=wb.active; info.title="Общая информация"
rows=[("Название компании","ООО «ПримерСерт» (демо)"),("ИНН","7700000000"),
      ("Город базирования","Москва"),("География продаж","Вся РФ"),
      ("Основные виды деятельности","Сертификация, экология, обучение (демо)"),
      ("Сайт компании","example.ru"),
      ("Основные преимущества","1. Скорость\n2. Репутация\n3. Финансовая стабильность\n4. Цена"),
      ("Основные слабые места","1. Бюрократия\n2. Слабый маркетинг\n3. Нет зрелого бренда"),
      ("Целевые клиенты","Промышленные предприятия"),
      ("Какие каналы продаж","1. База\n2. Звонки\n3. Тендеры")]
for i,(k,v) in enumerate(rows,1): info.cell(i,1,k); info.cell(i,2,v)

SHEETS={"Экологические услуги":[("Экологический аудит","20 000 – 300 000","2 нед – 3 мес","1 раз в 1–3 года","Нет","Гл. эколог","Конкурент А, Конкурент Б",4),
                               ("Экологическая отчётность","4 000 – 50 000","2–4 недели","Ежеквартально","Сильная (апрель, июль)","Гл. бухгалтер","Конкурент А, Конкурент В",2),
                               ("ПЭК","20 000 – 75 000","2–6 недель","Разовая","Нет","Гл. эколог","Конкурент Б",2)],
        "Сертификация продукции":[("Сертификация ТР ТС","120 000 – 350 000","2–4 месяца","1 раз в 5 лет","Нет","Гл. инженер","Конкурент Г, Конкурент Д",5),
                                  ("Декларация соответствия","7 000 – 50 000","2–5 дней","1 раз в 1–5 лет","Слабая","Рук. отдела","Конкурент Д",1)],
        "Учебный центр":[("Повышение квалификации","25 000 – 45 000","1–2 недели","1 раз в 3–5 лет","Не выражена","Рук. отдела","Конкурент Е, Конкурент Ж",2),
                         ("Корпоративное обучение","80 000 – 350 000","1–3 месяца","По запросу","Слабая (осень)","HR-директор","Конкурент Ж",4)],
        "Сертификация СМК":[("Сертификация ISO 9001","150 000 – 500 000","1–3 месяца","Ежегодно","Не выражена","Директор по качеству","Конкурент З, Конкурент И",3),
                            ("Консалтинг и аудит СМ","450 000 – 850 000","3–8 месяцев","1 раз в 3 года","Слабая","Ген. директор","Конкурент З",5)]}
for sn, prods in SHEETS.items():
    ws=wb.create_sheet(sn); ws.cell(1,1,"Название отдела"); ws.cell(1,2,sn)
    for c,h in enumerate(["","Услуга","Средний чек (руб.)","Цикл сделки","Повторяемость","Сезонность","ЛПР","Основные конкуренты","Сложность продажи"],1): ws.cell(3,c,h)
    r=4
    for p in prods:
        ws.cell(r,2,p[0]); ws.cell(r,3,p[1]); ws.cell(r,4,p[2]); ws.cell(r,5,p[3]); ws.cell(r,6,p[4]); ws.cell(r,7,p[5]); ws.cell(r,8,p[6]); ws.cell(r,9,p[7]); r+=1
wb.save(os.path.join(HERE,"zero_layer.xlsx")); print("xlsx -> zero_layer.xlsx")
print("ГОТОВО: синтетика в", HERE)
